import openpyxl


# wb = openpyxl.load_workbook(r"D:\AutomationTesting\PyCharmWS\nopCommerce\TestData\TestData.xlsx")
# ws = wb["Sheet1"]
# for row in ws.iter_rows():
#     print(row[0].value)
#     # for i in row:
#     #     print(i.value)
#
# for i in range(1, ws.max_row+1):
#     cellValue = [Cell.value for Cell in ws[i]]
#     print(cellValue)


def getRowCount(file, sheetname):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetname]
    return ws.max_row


def getColumnCount(file, sheetname):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetname]
    return ws.max_column


def readExcelData(file, sheetname, rownum, colnum):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetname]
    return ws.cell(row=rownum, column=colnum).value
